from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, flash, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
import PyPDF2
import openpyxl
from openpyxl import Workbook
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import json
import os.path
from datetime import datetime
import sqlite3
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
import uuid

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.secret_key = 'your-secret-key-here'  # Change this in production
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///recruitment.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db = SQLAlchemy(app)
csrf = CSRFProtect(app)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'hr' or 'candidate'
    name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Job(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    requirements = db.Column(db.Text)
    location = db.Column(db.String(100))
    company = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Application(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    resume_path = db.Column(db.String(500), nullable=False)
    cover_letter = db.Column(db.Text)
    status = db.Column(db.String(20), default='Pending')  # Pending, Reviewed, Accepted, Rejected
    applied_date = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Application details
    score = db.Column(db.Float, nullable=True)  # Matching score (0-100)
    
    # Relationships
    candidate = db.relationship('User', backref='applications')
    job = db.relationship('Job')

# Create database tables
with app.app_context():
    db.create_all()

# Dummy HR user (for demo purposes only)
with app.app_context():
    if not User.query.filter_by(username='hr101').first():
        hr_user = User(
            username='hr101',
            email='hr@example.com',
            password=generate_password_hash('hr1234'),
            role='hr',
            name='HR Manager'
        )
        db.session.add(hr_user)
        db.session.commit()

    # Add some sample jobs if none exist
    if Job.query.count() == 0:
        jobs = [
            Job(
                title='Senior Software Engineer',
                company='Tech Corp',
                location='New York, NY',
                description='We are looking for an experienced software engineer...',
                requirements='5+ years of experience with Python and web development...'
            ),
            Job(
                title='Product Manager',
                company='Innovate Inc',
                location='Remote',
                description='Lead our product development team...',
                requirements='3+ years of product management experience...'
            )
        ]
        db.session.bulk_save_objects(jobs)
        db.session.commit()

# Google Drive API Setup
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

def get_google_drive_service():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    
    return build('drive', 'v3', credentials=creds)

def extract_text_from_pdf(pdf_file):
    reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text

def screen_resume(resume_text, job_description):
    # Simple keyword matching for demo
    # In production, you might want to use NLP or more sophisticated matching
    keywords = job_description.lower().split()
    resume_words = resume_text.lower().split()
    
    # Simple scoring: percentage of keywords found in resume
    matched = sum(1 for word in keywords if word in resume_words)
    score = (matched / len(keywords)) * 100 if keywords else 0
    return min(100, round(score, 2))

def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('role_selection', next=request.url))
            
            user = User.query.get(session['user_id'])
            if not user:
                session.pop('user_id', None)
                return redirect(url_for('role_selection', next=request.url))
                
            if role and user.role != role:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for(f'{user.role}_dashboard'))
                
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Routes
@app.route('/')
def role_selection():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user.role == 'hr':
            return redirect(url_for('hr_dashboard'))
        else:
            return redirect(url_for('candidate_dashboard'))
    return render_template('role_selection.html')

@app.route('/login/<role>', methods=['GET', 'POST'])
def login(role):
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username, role=role).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            next_page = request.args.get('next')
            
            if role == 'hr':
                return redirect(next_page or url_for('hr_dashboard'))
            else:
                return redirect(next_page or url_for('candidate_dashboard'))
        else:
            flash('Invalid credentials. Please try again.', 'danger')
    
    if role == 'hr':
        return render_template('hr_login.html')
    else:
        return render_template('candidate_login.html', role=role)

@app.route('/hr/dashboard')
@login_required(role='hr')
def hr_dashboard():
    # Load resume database for HR view
    resume_db_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resume_database.json')
    resume_database = []
    if os.path.exists(resume_db_path):
        try:
            with open(resume_db_path, 'r') as f:
                resume_database = json.load(f)
        except Exception:
            resume_database = []

    # Add default score if not present
    for app in resume_database:
        if 'score' not in app:
            app['score'] = 0  # Default score if not present

    # Sort by date added (desc) if present
    def _parse_date(entry):
        try:
            return datetime.fromisoformat(entry.get('appliedDate') or entry.get('dateAdded', ''))
        except Exception:
            return datetime.min

    recent_applications = sorted(resume_database, key=_parse_date, reverse=True)[:20]

    return render_template('index.html', 
                         recent_applications=recent_applications,
                         score_num=0)  # Default score for the template

@app.route('/candidate/dashboard')
@login_required(role='candidate')
def candidate_dashboard():
    user = User.query.get(session['user_id'])
    
    # Load resume database
    resume_db_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resume_database.json')
    resume_database = []
    if os.path.exists(resume_db_path):
        with open(resume_db_path, 'r') as f:
            resume_database = json.load(f)
    
    # Filter applications for current user
    user_applications = [app for app in resume_database if app.get('email') == user.email]
    
    # Get available jobs
    available_jobs = Job.query.filter_by(is_active=True).all()
    
    return render_template('candidate_dashboard.html', 
                         current_user=user,
                         applications=user_applications,
                         available_jobs=available_jobs)

@app.route('/candidate/apply', methods=['POST'])
@login_required(role='candidate')
def submit_application():
    if 'resume' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(request.url)
    
    resume = request.files['resume']
    if resume.filename == '':
        flash('No selected file', 'danger')
        return redirect(request.url)
    
    if resume:
        # Get job details
        job_id = request.form.get('job_id')
        job = Job.query.get(job_id)
        if not job:
            flash('Invalid job selected', 'danger')
            return redirect(url_for('candidate_dashboard'))
        
        # Save the resume file
        filename = secure_filename(f"{session['user_id']}_{int(datetime.now().timestamp())}_{resume.filename}")
        resume_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resumes', filename)
        os.makedirs(os.path.dirname(resume_path), exist_ok=True)
        resume.save(resume_path)
        
        # Create application record in database
        application = Application(
            candidate_id=session['user_id'],
            job_id=job_id,
            resume_path=resume_path,
            cover_letter=request.form.get('cover_letter'),
            status='Pending'
        )
        db.session.add(application)
        db.session.commit()
        
        # Also add to resume database
        resume_db_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resume_database.json')
        resume_data = {
            'id': str(uuid.uuid4()),
            'name': request.form.get('full_name', ''),
            'email': request.form.get('email', ''),
            'phone': request.form.get('phone', ''),
            'position': job.title,
            'company': job.company,
            'dateAdded': datetime.now().strftime('%Y-%m-%d'),
            'status': 'New',
            'source': 'Portal',
            'fileName': filename,
            'filePath': resume_path,
            'applicationId': application.id,
            'appliedDate': datetime.now().isoformat(),
            'jobId': job_id
        }
        
        # Load existing data and append new entry
        existing_data = []
        if os.path.exists(resume_db_path):
            with open(resume_db_path, 'r') as f:
                existing_data = json.load(f)
        
        existing_data.append(resume_data)
        
        # Save back to file
        with open(resume_db_path, 'w') as f:
            json.dump(existing_data, f, indent=2)
        
        flash('Application submitted successfully!', 'success')
        return redirect(url_for('candidate_dashboard'))
    
    flash('Error processing your application', 'danger')
    return redirect(url_for('candidate_dashboard'))

@app.route('/candidate/upload-resume', methods=['POST'])
@login_required(role='candidate')
def upload_resume():
    if 'resume' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('candidate_dashboard'))
    
    resume = request.files['resume']
    if resume.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('candidate_dashboard'))
    
    if resume:
        # Save the resume file
        filename = secure_filename(f"{session['user_id']}_{int(datetime.now().timestamp())}_{resume.filename}")
        resume_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resumes', filename)
        os.makedirs(os.path.dirname(resume_path), exist_ok=True)
        resume.save(resume_path)
        
        # Update user's resume path (or create a new resume record)
        # This is a simplified example - you might want to create a separate Resume model
        
        flash('Resume uploaded successfully!', 'success')
    
    return redirect(url_for('candidate_dashboard'))

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('role_selection'))

@app.route('/upload_jd', methods=['POST'])
@login_required
def upload_jd():
    if 'job_description' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['job_description']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'jd.pdf')
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        file.save(filepath)
        return jsonify({'message': 'Job description uploaded successfully'})

@app.route('/screen_resumes', methods=['POST'])
@login_required(role='hr')
def screen_resumes():
    try:
        # Get job description
        jd_path = os.path.join(app.config['UPLOAD_FOLDER'], 'jd.pdf')
        if not os.path.exists(jd_path):
            return jsonify({'error': 'Please upload a job description first'}), 400
            
        with open(jd_path, 'rb') as jd_file:
            job_description = extract_text_from_pdf(jd_file)
        
        # Get resumes from uploads folder
        resumes_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'resumes')
        if not os.path.exists(resumes_dir):
            return jsonify({'error': 'No resumes found. Please upload some resumes first.'}), 400
            
        # Get list of resume files
        resume_files = [f for f in os.listdir(resumes_dir) if f.lower().endswith(('.pdf', '.doc', '.docx'))]
        
        if not resume_files:
            return jsonify({'error': 'No valid resume files found. Please upload PDF or Word documents.'}), 400
        
        # Process each resume
        results = []
        for resume_file in resume_files:
            try:
                resume_path = os.path.join(resumes_dir, resume_file)
                with open(resume_path, 'rb') as f:
                    resume_text = extract_text_from_pdf(f) if resume_file.lower().endswith('.pdf') else ""
                    # For non-PDF files, you might want to add text extraction for .doc/.docx
                    if not resume_text and resume_file.lower().endswith(('.doc', '.docx')):
                        resume_text = f"[Content from {resume_file} - text extraction for .doc/.docx not implemented]"
                    
                    # Simple scoring based on keyword matching
                    score = screen_resume(resume_text, job_description)
                    
                    # Get candidate info from filename or database
                    # For now, we'll use the filename
                    candidate_name = ' '.join(resume_file.split('_')[:-2])  # Remove timestamp and extension
                    
                    results.append({
                        'name': candidate_name,
                        'email': f"{candidate_name.replace(' ', '.').lower()}@example.com",
                        'score': score,
                        'status': 'Pending Review',
                        'resume': resume_file
                    })
            except Exception as e:
                print(f"Error processing {resume_file}: {str(e)}")
                continue
        
        # Sort results by score (highest first)
        results.sort(key=lambda x: x['score'], reverse=True)
        
        # Save results to session for download
        session['screening_results'] = results
        
        # Save to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Screened Candidates"
        
        # Add headers
        headers = ['Name', 'Email', 'Score (%)']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=result['name'])
            ws.cell(row=row_num, column=2, value=result['email'])
            ws.cell(row=row_num, column=3, value=result['score'])
        
        # Save the workbook
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'screened_candidates.xlsx')
        wb.save(excel_path)
        
        return jsonify({
            'message': f'Successfully screened {len(results)} resumes',
            'results': results,
            'excel_path': excel_path
        })
        
    except Exception as e:
        return jsonify({'error': f'Error screening resumes: {str(e)}'}), 500

@app.route('/download_results')
@login_required(role='hr')
def download_results():
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'screened_candidates.xlsx')
    if not os.path.exists(excel_path):
        return jsonify({'error': 'No results available'}), 404
    
    return send_file(
        excel_path,
        as_attachment=True,
        download_name='screened_candidates.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True)
